import io
import openpyxl
from fastapi import HTTPException


REQUIRED_COLUMNS = {"firstname", "lastname"}
OPTIONAL_COLUMNS = {"othername", "email", "phone", "institution_id", "access_code"}


def generate_student_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["firstname", "lastname", "othername", "email", "phone", "institution_id", "access_code"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4A90E2")

    # Example rows
    examples = [
        ["John", "Doe", "Michael", "john@school.com", "+2348012345678", "STU001", ""],
        ["Jane", "Smith", "", "", "+2348087654321", "STU002", "ABC123"],
        ["Bob", "Johnson", "", "bob@school.com", "", "STU003", ""],
    ]
    for row in examples:
        ws.append(row)

    # Instructions sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ["Column", "Required", "Description"],
        ["firstname", "Yes", "Student's first name"],
        ["lastname", "Yes", "Student's last name"],
        ["othername", "No", "Middle name or other name"],
        ["email", "No", "Student email — if provided, access code will be emailed"],
        ["phone", "No", "Phone number with country code e.g. +2348012345678"],
        ["institution_id", "No", "Student registration/admission number — must be unique per organization"],
        ["access_code", "No", "Custom access code — leave blank to auto-generate a unique one"],
    ]
    for row in instructions:
        ws_info.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def parse_student_excel(file_bytes: bytes, filename: str) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file. Please use the provided template.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File is empty or missing data rows.")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}. Please use the provided template."
        )

    parsed = []
    for row_data in rows[1:]:
        row = {
            k: (str(v).strip() if v is not None else "")
            for k, v in zip(headers, row_data)
        }
        if any(row.values()):  # skip completely empty rows
            parsed.append(row)

    return parsed