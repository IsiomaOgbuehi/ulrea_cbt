import pytest
import io
import openpyxl
from .conftest import create_subject, assign_teacher, TEACHER_ID


MCQ_ITEM = {
    "question_text": "What is 2 + 2?",
    "item_type": "mcq_single",
    "options": [
        {"key": "A", "text": "3"},
        {"key": "B", "text": "4"},
        {"key": "C", "text": "5"},
        {"key": "D", "text": "6"},
    ],
    "correct_answers": ["B"],
    "explanation": "Basic arithmetic.",
    "marks": 1.0,
    "negative_marks": 0.0,
    "tags": ["math"],
    "difficulty": "easy",
}

SHORT_ANSWER_ITEM = {
    "question_text": "Describe the water cycle.",
    "item_type": "short_answer",
    "marks": 5.0,
}


# ============================================================
# ITEM CRUD
# ============================================================

def test_teacher_can_create_item(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=MCQ_ITEM,
        headers=teacher_headers,
    )
    assert response.status_code == 200, response.json()
    data = response.json()
    assert data["question_text"] == MCQ_ITEM["question_text"]
    assert data["item_type"] == "mcq_single"
    assert data["status"] == "active"
    assert data["version"] == 1
    assert data["source"] == "manual"


def test_teacher_cannot_create_item_in_unassigned_subject(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Chemistry")
    # No assignment

    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=MCQ_ITEM,
        headers=teacher_headers,
    )
    assert response.status_code == 403


def test_mcq_requires_options(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    bad_item = {**MCQ_ITEM, "options": ["only one option"]}
    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=bad_item,
        headers=teacher_headers,
    )
    assert response.status_code == 422


def test_mcq_requires_correct_answer(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    bad_item = {**MCQ_ITEM, "correct_answers": None}
    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=bad_item,
        headers=teacher_headers,
    )
    assert response.status_code == 422


def test_short_answer_needs_no_options(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "English")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=SHORT_ANSWER_ITEM,
        headers=teacher_headers,
    )
    assert response.status_code == 200


def test_list_items_filtered_by_difficulty(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    client.post(f"/api/v1/subjects/{subject['id']}/items", json={**MCQ_ITEM, "difficulty": "easy"}, headers=teacher_headers)
    client.post(f"/api/v1/subjects/{subject['id']}/items", json={**MCQ_ITEM, "question_text": "Hard Q", "difficulty": "hard"}, headers=teacher_headers)

    response = client.get(
        f"/api/v1/subjects/{subject['id']}/items?difficulty=easy",
        headers=teacher_headers,
    )
    assert response.status_code == 200
    assert all(i["difficulty"] == "easy" for i in response.json())


def test_list_items_search(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    client.post(f"/api/v1/subjects/{subject['id']}/items", json={**MCQ_ITEM, "question_text": "What is gravity?"}, headers=teacher_headers)
    client.post(f"/api/v1/subjects/{subject['id']}/items", json={**MCQ_ITEM, "question_text": "What is velocity?"}, headers=teacher_headers)

    response = client.get(
        f"/api/v1/subjects/{subject['id']}/items?search=gravity",
        headers=teacher_headers,
    )
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert "gravity" in data[0]["question_text"]


def test_update_item_increments_version(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    create_response = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=MCQ_ITEM,
        headers=teacher_headers,
    ).json()

    update_response = client.patch(
        f"/api/v1/subjects/{subject['id']}/items/{create_response['id']}",
        json={"question_text": "Updated question_text"},
        headers=teacher_headers,
    )
    assert update_response.status_code == 200
    assert update_response.json()["version"] == 2
    assert update_response.json()["question_text"] == "Updated question_text"


def test_delete_item_archives_it(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    item = client.post(
        f"/api/v1/subjects/{subject['id']}/items",
        json=MCQ_ITEM,
        headers=teacher_headers,
    ).json()

    delete_response = client.delete(
        f"/api/v1/subjects/{subject['id']}/items/{item['id']}",
        headers=teacher_headers,
    )
    assert delete_response.status_code == 204

    # Archived item not in active list
    list_response = client.get(
        f"/api/v1/subjects/{subject['id']}/items?status=active",
        headers=teacher_headers,
    )
    ids = [i["id"] for i in list_response.json()]
    assert item["id"] not in ids


# ============================================================
# BULK UPLOAD
# ============================================================

def make_excel(rows: list[list]) -> bytes:
    """Helper to create an in-memory Excel file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "question_text", "item_type", "option_a", "option_b", "option_c", "option_d",
        "correct_answers", "explanation", "marks", "negative_marks", "tags", "difficulty"
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def test_bulk_upload_valid_rows(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    rows = [
        ["What is 2+2?", "mcq_single", "3", "4", "5", "6", "B", "Arithmetic", 1, 0, "math", "easy"],
        ["Which are prime?", "mcq_multi", "2", "3", "4", "5", "A,B,D", "", 2, 0.5, "math", "medium"],
        ["Sky is blue.", "true_false", "True", "False", "", "", "True", "", 1, 0, "", "easy"],
        ["Speed of light (m/s)?", "numeric", "", "", "", "", "299792458", "", 2, 0, "physics", "hard"],
    ]
    file_bytes = make_excel(rows)

    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items/bulk",
        headers=teacher_headers,
        files={"file": ("questions.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.json()
    data = response.json()
    assert data["total_rows"] == 4
    assert data["successful_rows"] == 4
    assert data["failed_rows"] == 0
    assert "upload_id" in data


def test_bulk_upload_partial_failure(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Science")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    rows = [
        ["Valid question?", "mcq_single", "A", "B", "C", "D", "A", "", 1, 0, "", "easy"],
        ["", "mcq_single", "A", "B", "", "", "A", "", 1, 0, "", ""],      # missing question_text
        ["Bad type?", "invalid_type", "A", "B", "", "", "A", "", 1, 0, "", ""],  # bad type
    ]
    file_bytes = make_excel(rows)

    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items/bulk",
        headers=teacher_headers,
        files={"file": ("questions.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["successful_rows"] == 1
    assert data["failed_rows"] == 2
    assert len(data["errors"]) == 2


def test_bulk_upload_wrong_file_type(client, admin_headers, teacher_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    response = client.post(
        f"/api/v1/subjects/{subject['id']}/items/bulk",
        headers=teacher_headers,
        files={"file": ("questions.csv", b"question_text,item_type\nQ1,mcq_single", "text/csv")},
    )
    assert response.status_code == 400
    assert "xlsx" in response.json()["detail"].lower()


def test_download_template(client, teacher_headers, admin_headers):
    subject = create_subject(client, admin_headers, "Math")
    assign_teacher(client, admin_headers, subject["id"], TEACHER_ID)

    response = client.get(
        f"/api/v1/subjects/{subject['id']}/items/bulk/template",
        headers=teacher_headers,
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # Verify it's a valid Excel file
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Questions" in wb.sheetnames
    assert "Instructions" in wb.sheetnames
