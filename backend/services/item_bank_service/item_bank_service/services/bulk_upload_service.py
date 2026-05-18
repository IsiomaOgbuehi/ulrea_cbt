from uuid import UUID, uuid4
import io
import openpyxl
from fastapi import HTTPException
from sqlmodel import Session

from item_bank_service.database.models.item import ItemModel, BulkUploadLog
from item_bank_service.database.models.enums import ItemType, ItemSource
from item_bank_service.schemas.schemas import BulkUploadResult, CurrentUser


# ============================================================
# CONSTANTS
# ============================================================

REQUIRED_COLUMNS = {"question_text", "item_type"}

OPTIONAL_COLUMNS = {
    "option_a", "option_b", "option_c", "option_d", "option_e",
    "correct_answers", "explanation", "marks", "negative_marks",
    "tags", "difficulty"
}

VALID_TYPES = {t.value for t in ItemType}
VALID_DIFFICULTIES = {"easy", "medium", "hard"}


# ============================================================
# SAFE UTIL
# ============================================================

def _safe_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


# ============================================================
# PARSERS
# ============================================================

def _parse_options(row: dict) -> list[dict] | None:
    """
    Returns JSON-safe list for ItemOption:
    [
        {"key": "A", "text": "Bike"}
    ]
    """

    mapping = [
        ("A", "option_a"),
        ("B", "option_b"),
        ("C", "option_c"),
        ("D", "option_d"),
        ("E", "option_e"),
    ]

    options = []

    for key, col in mapping:
        val = _safe_str(row.get(col)).strip()
        if val:
            options.append({"key": key, "text": val})

    return options or None


def _parse_correct_answers(raw: str) -> list[str] | None:
    raw = _safe_str(raw).strip()
    if not raw:
        return None

    return [x.strip() for x in raw.split(",") if x.strip()]


def _parse_tags(raw: str) -> list[str] | None:
    raw = _safe_str(raw).strip()
    if not raw:
        return None

    return [t.strip() for t in raw.split(",") if t.strip()]


# ============================================================
# VALIDATION
# ============================================================

def _validate_row(row: dict, row_num: int):
    question_text = _safe_str(row.get("question_text")).strip()
    if not question_text:
        return None, f"Row {row_num}: question_text is required"

    raw_type = _safe_str(row.get("item_type")).strip().lower()

    if raw_type not in VALID_TYPES:
        return None, f"Row {row_num}: invalid item_type '{raw_type}'"

    options = _parse_options(row)

    raw_correct = row.get("correct_answers") or row.get("correct_answer")
    correct_answers = _parse_correct_answers(raw_correct)

    difficulty = _safe_str(row.get("difficulty")).strip().lower() or None
    if difficulty and difficulty not in VALID_DIFFICULTIES:
        return None, f"Row {row_num}: invalid difficulty '{difficulty}'"

    try:
        marks = float(row.get("marks") or 1.0)
        negative_marks = float(row.get("negative_marks") or 0.0)
    except Exception:
        return None, f"Row {row_num}: marks must be numeric"

    # RULES
    if raw_type in (ItemType.MCQ_SINGLE.value, ItemType.MCQ_MULTI.value, ItemType.TRUE_FALSE.value):
        if not options or len(options) < 2:
            return None, f"Row {row_num}: MCQ/TF requires at least 2 options"
        if not correct_answers:
            return None, f"Row {row_num}: correct_answers required"

    if raw_type == ItemType.NUMERIC.value and not correct_answers:
        return None, f"Row {row_num}: numeric requires correct_answers"

    return {
        "question_text": question_text,
        "item_type": raw_type,
        "options": options,
        "correct_answers": correct_answers,
        "explanation": _safe_str(row.get("explanation")).strip() or None,
        "marks": marks,
        "negative_marks": negative_marks,
        "tags": _parse_tags(row.get("tags")),
        "difficulty": difficulty,
    }, None


# ============================================================
# SERVICE
# ============================================================

class BulkUploadService:

    # --------------------------------------------------------
    # TEMPLATE
    # --------------------------------------------------------
    @staticmethod
    def generate_template() -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"

        headers = [
            "question_text", "item_type",
            "option_a", "option_b", "option_c", "option_d", "option_e",
            "correct_answers", "explanation",
            "marks", "negative_marks", "tags", "difficulty"
        ]
        ws.append(headers)

        examples = [
            ["What is 2+2?", "mcq_single", "3", "4", "5", "6", "", "B", "Basic math", 1, 0, "math", "easy"],
            ["Prime numbers?", "mcq_multi", "2", "3", "4", "5", "", "A,B,D", "", 2, 0.5, "math", "medium"],
            ["Sky is blue", "true_false", "True", "False", "", "", "", "True", "", 1, 0, "general", "easy"],
            ["Speed of light?", "numeric", "", "", "", "", "", "299792458", "", 2, 0, "physics", "hard"],
            ["Describe photosynthesis", "short_answer", "", "", "", "", "", "", "Manual review", 5, 0, "biology", "medium"],
        ]
        for r in examples:
            ws.append(r)

        # Instructions sheet
        ws_info = wb.create_sheet("Instructions")
        instructions = [
            ["Column", "Required", "Description"],
            ["question_text", "Yes", "The question text"],
            ["item_type", "Yes", "mcq_single | mcq_multi | true_false | numeric | short_answer"],
            ["option_a to option_e", "For MCQ/TF", "Answer options (at least 2 for MCQ)"],
            ["correct_answers", "For most types", "Option key (A, B...). Multi: comma-separated (A,C). Numeric: the number."],
            ["explanation", "No", "Shown to candidate after exam"],
            ["marks", "No", "Points awarded. Default: 1"],
            ["negative_marks", "No", "Points deducted for wrong answer. Default: 0"],
            ["tags", "No", "Comma-separated tags e.g. algebra,chapter-3"],
            ["difficulty", "No", "easy | medium | hard"],
        ]
        for row in instructions:
            ws_info.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # --------------------------------------------------------
    # PROCESS UPLOAD
    # --------------------------------------------------------
    @staticmethod
    def process_upload(
        session: Session,
        file_bytes: bytes,
        filename: str,
        subject_id: UUID,
        current_user: CurrentUser,
    ) -> BulkUploadResult:

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid Excel file")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Empty file")

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing columns: {', '.join(missing)}"
            )

        upload_id = uuid4()
        successful = []
        errors = []

        for row_num, row_data in enumerate(rows[1:], start=2):

            row = {
                k: (str(v).strip() if v is not None else "")
                for k, v in zip(headers, row_data)
            }

            if not any(row.values()):
                continue

            cleaned, error = _validate_row(row, row_num)

            if error:
                errors.append({"row": row_num, "error": error})
                continue

            item = ItemModel(
                org_id=current_user.org_id,
                subject_id=subject_id,
                created_by=current_user.id,
                source=ItemSource.EXCEL_UPLOAD,
                bulk_upload_id=upload_id,
                **cleaned,
            )

            session.add(item)
            successful.append(item)

        if successful:
            session.commit()

        log = BulkUploadLog(
            id=upload_id,
            org_id=current_user.org_id,
            subject_id=subject_id,
            uploaded_by=current_user.id,
            filename=filename,
            total_rows=len(rows) - 1,
            successful_rows=len(successful),
            failed_rows=len(errors),
            errors=errors,
        )

        session.add(log)
        session.commit()

        return BulkUploadResult(
            total_rows=len(rows) - 1,
            successful_rows=len(successful),
            failed_rows=len(errors),
            errors=errors,
            upload_id=upload_id,
        )